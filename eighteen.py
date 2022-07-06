import pandas as pd
import random
from openpyxl.styles import Border, Side
import openpyxl


def question():
    type_of_problem = random.randint(1, 4)
    if type_of_problem == 1:
        aaa = random.randint(1, 3)
        if aaa == 1:
            c_a = top_left_easy()
        elif aaa == 2:
            c_a = top_left_medium()
        else:
            c_a = top_left_hard()
    elif type_of_problem == 2:
        aaa = random.randint(1, 3)
        if aaa == 1:
            c_a = top_right()
        elif aaa == 2:
            c_a = top_right_M()
        else:
            c_a = top_right_hard()
    elif type_of_problem == 3:
        aaa = random.randint(1, 3)
        if aaa == 1:
            c_a = bottom_left()
        elif aaa == 2:
            c_a = bottom_left_M()
        else:
            c_a = bottom_left_hard()
    else:
        aaa = random.randint(1, 3)
        if aaa == 1:
            c_a = bottom_right()
        elif aaa == 2:
            c_a = bottom_right_M()
        else:
            c_a = bottom_right_hard()
    return c_a


def create_main_table(a_o_l):
    sol_table = []
    for i in range(a_o_l):
        line = []
        for j in range(a_o_l):
            line.append(0)
        sol_table.append(line)
    return sol_table


def start(amount_of_lines, smth_1, smth_2, smth_3, smth_4):
    print('Квадрат разлинован на {}*{} клеток.'.format(amount_of_lines, amount_of_lines))
    print(
        'Исполнитель Робот может перемещаться по клеткам, '
        'выполняя за одно перемещение одну из двух команд: {} или {}.'.format(
            smth_1, smth_2))
    print('По команде {} Робот перемещается в соседнюю {}'
          ' клетку, по команде {} — в соседнюю {}.'.format(smth_1, smth_3, smth_2, smth_4))
    print('При попытке выхода за границу квадрата Робот разрушается.')
    print('Так же в файле могут встречаться стенки, ударяясь об которые робот ломается')
    print('Перед каждым запуском Робота в каждой клетке квадрата лежит монета достоинством от 1 до 550.')


def gen_file():
    x = random.randint(11, 19)
    first_table = []
    sex_after_siggarets = ''
    for i in range(x):  # Number
        string = ''
        for j in range(x):  # Letter
            smth = random.randint(1, 550)
            while str(smth) in sex_after_siggarets:
                smth = random.randint(1, 550)
            string += str(smth) + ' '
            sex_after_siggarets += str(smth) + ' '
        strings = string.split()
        strings = list(map(int, strings))
        first_table.append(strings)
    dictionary = dict()
    for i in range(x):  # Number
        for j in range(x):  # Letter
            i_hate_my_life = first_table[i][j]
            key_from_dict = first_table[0][j]
            if i == 0:
                dictionary[i_hate_my_life] = [i_hate_my_life]
            else:
                dictionary[key_from_dict].append(i_hate_my_life)
    df = pd.DataFrame(dictionary)
    df.to_excel('18.xlsx', index=False, header=False)
    wb = openpyxl.load_workbook('18.xlsx')
    ws = wb['Sheet1']
    type_and_color_borders = Side(border_style='thick', color='000000')

    b_b = Border(bottom=type_and_color_borders)
    b_r = Border(right=type_and_color_borders)
    letter = chr(64 + x)
    number = x
    start_cell = letter + str(number)
    for smth in range(1, x + 1):
        cell_m_n = letter + str(smth)
        cell_m_L = chr(64 + smth) + str(number)
        ws[cell_m_n].border = b_r
        ws[cell_m_L].border = b_b
    ws[start_cell].border = Border(right=type_and_color_borders, bottom=type_and_color_borders)
    wb.save('18.xlsx')
    return first_table, x


def example():
    print('Пример входных данных:')
    a = [[1, 8, 8, 4], [10, 1, 1, 3], [1, 3, 12, 2], [2, 3, 5, 6]]
    for i in a:
        print(*i)


def example_solution(psih, type_position):
    a = [[1, 8, 8, 4], [10, 1, 1, 3], [1, 3, 12, 2], [2, 3, 5, 6]]
    if type_position == 1:
        a.reverse()  # bottom left
    elif type_position == 2:  # bottom right
        a.reverse()
        for smth in range(len(a)):
            a[smth].reverse()
    elif type_position == 3:  # top right
        for smth in range(len(a)):
            a[smth].reverse()
    answer = []
    solution_table = create_main_table(4)

    if a[0][0] % psih == 0:
        solution_table[0][0] = a[0][0]
    else:
        solution_table[0][0] = 0

    for i in range(1, 4):
        if a[i][0] % psih == 0:
            solution_table[i][0] = solution_table[i - 1][0] + a[i][0]
        else:
            solution_table[i][0] = solution_table[i - 1][0]
        if a[0][i] % psih == 0:
            solution_table[0][i] = solution_table[0][i - 1] + a[0][i]
        else:
            solution_table[0][i] = solution_table[0][i - 1]

    for i in range(1, 4):
        for j in range(1, 4):
            if a[i][j] % psih == 0:
                solution_table[i][j] = max(solution_table[i - 1][j], solution_table[i][j - 1]) + a[i][j]
            else:
                solution_table[i][j] = max(solution_table[i - 1][j], solution_table[i][j - 1])
    answer.append(solution_table[-1][-1])
    for i in range(1, 4):
        for j in range(1, 4):
            if a[i][j] % psih == 0:
                solution_table[i][j] = min(solution_table[i - 1][j], solution_table[i][j - 1]) + a[i][j]
            else:
                solution_table[i][j] = min(solution_table[i - 1][j], solution_table[i][j - 1])
    answer.append(solution_table[-1][-1])
    return answer


def solution_easy(table: 'input table from EXCEL', a_o_l: 'Amount of lines'):
    solution_table = create_main_table(
        a_o_l)  # Table with full solution of this problem like a solution table[NUMBER][LETTER]

    for i in range(1, a_o_l):
        solution_table[i][0] = solution_table[i - 1][0] + table[i][0]
        solution_table[0][i] = solution_table[0][i - 1] + table[0][i]

    def resolving(amount_of_lines, return_table, main_table, value):
        for f in range(1, amount_of_lines):
            for j in range(1, amount_of_lines):
                if value == 1:
                    return_table[f][j] = max(return_table[f - 1][j], return_table[f][j - 1]) + main_table[f][j]
                else:
                    return_table[f][j] = min(return_table[f - 1][j], return_table[f][j - 1]) + main_table[f][j]
        return return_table

    solution_table = resolving(a_o_l, solution_table, table, 1)

    maximum = solution_table[a_o_l - 1][a_o_l - 1]  # First answer in our problem
    solution_table = resolving(a_o_l, solution_table, table, 2)

    minimum = solution_table[a_o_l - 1][a_o_l - 1]  # Second answer in our problem
    answer = [maximum, minimum]  # Full answer in one variable(probably changed)
    return answer


def top_left_easy():
    table, amount_of_lines = gen_file()  # Excel table, amount of lines and values in lines
    correct_answer = solution_easy(table, amount_of_lines)  # Correct answer taken from solution
    start(amount_of_lines, 'вниз', 'право', 'нижнюю', 'правую')
    print(
        'Посетив клетку, Робот забирает монету с собой; '
        'это также относится к начальной и конечной клетке маршрута Робота.')
    print(
        'Откройте файл. Определите максимальную и минимальную денежную сумму, '
        'которую может собрать Робот, пройдя из левой верхней клетки в правую нижнюю.')
    print(
        'В ответ запишите два числа друг за другом через пробел — сначала максимальную сумму, затем минимальную.')
    print(
        'Исходные данные представляют собой электронную таблицу размером {}*{}, '
        'каждая ячейка которой соответствует клетке квадрата.'.format(amount_of_lines, amount_of_lines))
    example()
    print('Для указанных входных данных ответом должна быть пара чисел 41 и 22.')
    return correct_answer


def solution_medium(table: ' Excel table', a_o_l, psih: ' Variable divider'):
    answer = []
    solution_table = create_main_table(
        a_o_l)  # Table with full solution of this problem like a solution table[NUMBER][LETTER]

    if table[0][0] % psih == 0:
        solution_table[0][0] = table[0][0]
    else:
        solution_table[0][0] = 0
    for i in range(1, a_o_l):
        if table[i][0] % psih == 0:
            solution_table[i][0] = solution_table[i - 1][0] + table[i][0]
        else:
            solution_table[i][0] = solution_table[i - 1][0]
        if table[0][i] % psih == 0:
            solution_table[0][i] = solution_table[0][i - 1] + table[0][i]
        else:
            solution_table[0][i] = solution_table[0][i - 1]

    for i in range(1, a_o_l):
        for j in range(1, a_o_l):
            if table[i][j] % psih == 0:
                solution_table[i][j] = max(solution_table[i - 1][j], solution_table[i][j - 1]) + table[i][j]
            else:
                solution_table[i][j] = max(solution_table[i - 1][j], solution_table[i][j - 1])

    answer.append(solution_table[-1][-1])  # Adding to correct answer first value for this problem
    for i in range(1, a_o_l):
        for j in range(1, a_o_l):
            if table[i][j] % psih == 0:
                solution_table[i][j] = min(solution_table[i - 1][j], solution_table[i][j - 1]) + table[i][j]
            else:
                solution_table[i][j] = min(solution_table[i - 1][j], solution_table[i][j - 1])

    answer.append(solution_table[-1][-1])  # Adding to correct answer seconde value for this problem
    return answer


def top_left_medium():
    table, amount_of_lines = gen_file()  # Excel table, amount of lines and values in lines
    psih = random.randint(2, 10)  # Random number that will be divider
    correct_answer = solution_medium(table, amount_of_lines, psih)
    start(amount_of_lines, 'вниз', 'право', 'нижнюю', 'правую')
    print(
        'Посетив клетку, Робот забирает монету с собой,если её достоинство кратно {};\n'
        ' это также относится к начальной и конечной клетке маршрута Робота.'.format(psih))
    print(
        'Откройте файл. Определите максимальную и минимальную денежную сумму, которую может собрать Робот, '
        'пройдя из левой верхней клетки в правую нижнюю.')
    print(
        'В ответ запишите два числа друг за другом через пробел — сначала максимальную сумму, затем минимальную.')
    print(
        'Исходные данные представляют собой электронную таблицу размером {}*{}, '
        'каждая ячейка которой соответствует клетке квадрата.'.format(amount_of_lines, amount_of_lines))
    example()
    c_e_1, c_e_2 = example_solution(psih, 0)  # Automatic answer for a table from funcution example
    print('Для указанных входных данных ответом должна быть пара чисел {} и {}.'.format(c_e_1, c_e_2))
    return correct_answer


def top_right_solution(table: 'input table from EXCEL', a_o_l: 'Amount of lines'):
    for x in range(len(table)):
        table[x].reverse()
    answer = solution_easy(table, a_o_l)
    return answer


def top_right():
    table, amount_of_lines = gen_file()
    correct_answer = top_right_solution(table, amount_of_lines)
    start(amount_of_lines, 'в низ', 'влево', 'нижнию', 'левую')
    print(
        'Посетив клетку, Робот забирает монету с собой; '
        'это также относится к начальной и конечной клетке маршрута Робота.')
    print(
        'Откройте файл. Определите максимальную и минимальную денежную сумму, '
        'которую может собрать Робот, пройдя из правой верхней в левую нижнюю.')
    print(
        'В ответ запишите два числа друг за другом через пробел — сначала максимальную сумму, затем минимальную.')
    print(
        'Исходные данные представляют собой электронную таблицу размером {}*{}, '
        'каждая ячейка которой соответствует клетке квадрата.'.format(amount_of_lines, amount_of_lines))
    example()
    print('Для указанных входных данных ответом должна быть пара чисел 35 и 15.')
    return correct_answer


def top_right_M_solution(table: ' Excel table', a_o_l, psih: ' Variable divider'):
    for x in range(len(table)):
        table[x].reverse()
    return solution_medium(table, a_o_l, psih)


def top_right_M():
    table, amount_of_lines = gen_file()  # Excel table, amount of lines and values in lines
    psih = random.randint(2, 10)  # Random number that will be divider
    correct_answer = top_right_M_solution(table, amount_of_lines, psih)
    start(amount_of_lines, 'в низ', 'влево', 'нижнию', 'левую')
    print(
        'Посетив клетку, Робот забирает монету с собой,если её достоинство кратно {};\n '
        'это также относится к начальной и конечной клетке маршрута Робота.'.format(
            psih))
    print(
        'Откройте файл. Определите максимальную и минимальную денежную сумму, '
        'которую может собрать Робот, пройдя из правой верхней клетки в левую нижнюю.')
    print(
        'В ответ запишите два числа друг за другом через пробел — сначала максимальную сумму, затем минимальную.')
    print(
        'Исходные данные представляют собой электронную таблицу размером {}*{},'
        ' каждая ячейка которой соответствует клетке квадрата.'.format(
            amount_of_lines, amount_of_lines))
    example()
    c_e_1, c_e_2 = example_solution(psih, 3)  # Automatic answer for a table from funcution example
    print('Для указанных входных данных ответом должна быть пара чисел {} и {}.'.format(c_e_1, c_e_2))
    return correct_answer


def bottom_left_soliution(table: 'input table from EXCEL', a_o_l: 'Amount of lines'):
    table.reverse()
    answer = solution_easy(table, a_o_l)
    return answer


def bottom_left():
    table, amount_of_lines = gen_file()
    correct_answer = bottom_left_soliution(table, amount_of_lines)
    start(amount_of_lines, 'вверх', 'право', 'верхнюю', 'правую')
    print(
        'Посетив клетку, Робот забирает монету с собой; '
        'это также относится к начальной и конечной клетке маршрута Робота.')
    print(
        'Откройте файл. Определите максимальную и минимальную денежную сумму, '
        'которую может собрать Робот, пройдя из левой нижней в правую верхнюю.')
    print(
        'В ответ запишите два числа друг за другом через пробел — сначала максимальную сумму, затем минимальную.')
    print(
        'Исходные данные представляют собой электронную таблицу размером {}*{}, '
        'каждая ячейка которой соответствует клетке квадрата.'.format(
            amount_of_lines, amount_of_lines))
    example()
    print('Для указанных входных данных ответом должна быть пара чисел 35 и 15.')
    return correct_answer


def bottom_left_M_solution(table: ' Excel table', a_o_l, psih: ' Variable divider'):
    table.reverse()
    return solution_medium(table, a_o_l, psih)


def bottom_left_M():
    table, amount_of_lines = gen_file()  # Excel table, amount of lines and values in lines
    psih = random.randint(2, 10)  # Random number that will be divider
    correct_answer = bottom_left_M_solution(table, amount_of_lines, psih)
    start(amount_of_lines, 'вверх', 'право', 'верхнюю', 'правую')
    print(
        'Посетив клетку, Робот забирает монету с собой,если её достоинство кратно {};\n'
        ' это также относится к начальной и конечной клетке маршрута Робота.'.format(
            psih))
    print(
        'Откройте файл. Определите максимальную и минимальную денежную сумму,'
        ' которую может собрать Робот, пройдя из левой нижней клетки в правую верхнюю.')
    print(
        'В ответ запишите два числа друг за другом через пробел — сначала максимальную сумму, затем минимальную.')
    print(
        'Исходные данные представляют собой электронную таблицу размером {}*{},'
        ' каждая ячейка которой соответствует клетке квадрата.'.format(
            amount_of_lines, amount_of_lines))
    example()
    c_e_1, c_e_2 = example_solution(psih, 1)  # Automatic answer for a table from funcution example
    print('Для указанных входных данных ответом должна быть пара чисел {} и {}.'.format(c_e_1, c_e_2))
    return correct_answer


def bottom_right_solution(table: 'input table from EXCEL', a_o_l: 'Amount of lines'):
    table.reverse()
    for x in range(len(table)):
        table[x].reverse()
    return solution_easy(table, a_o_l)


def bottom_right():
    table, amount_of_lines = gen_file()
    correct_answer = bottom_right_solution(table, amount_of_lines)
    start(amount_of_lines, 'вверх', 'влево', 'верхнюю', 'левую')
    print(
        'Посетив клетку, Робот забирает монету с собой; '
        'это также относится к начальной и конечной клетке маршрута Робота.')
    print(
        'Откройте файл. Определите максимальную и минимальную денежную сумму,'
        ' которую может собрать Робот, пройдя из правой нижней в левую верхнюю.')
    print(
        'В ответ запишите два числа друг за другом через пробел — сначала максимальную сумму, затем минимальную.')
    print(
        'Исходные данные представляют собой электронную таблицу размером {}*{},'
        ' каждая ячейка которой соответствует клетке квадрата.'.format(
            amount_of_lines, amount_of_lines))
    example()
    print('Для указанных входных данных ответом должна быть пара чисел 41 и 22.')
    return correct_answer


def bottom_right_M_solution(table: ' Excel table', a_o_l, psih: ' Variable divider'):
    table.reverse()
    for x in range(len(table)):
        table[x].reverse()
    return solution_medium(table, a_o_l, psih)


def bottom_right_M():
    table, amount_of_lines = gen_file()  # Excel table, amount of lines and values in lines
    psih = random.randint(2, 10)  # Random number that will be divider
    correct_answer = bottom_right_M_solution(table, amount_of_lines, psih)
    start(amount_of_lines, 'вверх', 'влево', 'верхнюю', 'левую')
    print(
        'Посетив клетку, Робот забирает монету с собой,если её достоинство кратно {};\n'
        ' это также относится к начальной и конечной клетке маршрута Робота.'.format(
            psih))
    print(
        'Откройте файл. Определите максимальную и минимальную денежную сумму,'
        ' которую может собрать Робот, пройдя из правой нижней клетки в левую верхнюю.')
    print(
        'В ответ запишите два числа друг за другом через пробел — сначала максимальную сумму, затем минимальную.')
    print(
        'Исходные данные представляют собой электронную таблицу размером {}*{},'
        ' каждая ячейка которой соответствует клетке квадрата.'.format(
            amount_of_lines, amount_of_lines))
    example()
    c_e_1, c_e_2 = example_solution(psih, 2)  # Automatic answer for a table from funcution example
    print('Для указанных входных данных ответом должна быть пара чисел {} и {}.'.format(c_e_1, c_e_2))
    return correct_answer


def gen_borders(L: str, corner_N, corner_L, len_N, len_L):
    top = bottom = right = left = Side(border_style='thick', color='000000')
    wb = openpyxl.load_workbook('18.xlsx')
    ws = wb['Sheet1']
    letter = chr(65 + corner_L)
    number = str(corner_N)
    corner = letter + number
    if L == 'TL':
        ws[corner].border = Border(bottom=bottom, right=right)
        for i in range(1, len_N):  # Moving by numbers
            cell = letter + str(int(number) - i)
            ws[cell].border = Border(right=right)
        for i in range(1, len_L):  # Moving by letters
            cell = chr(ord(letter) - i) + number
            ws[cell].border = Border(bottom=bottom)
    elif L == 'TR':
        ws[corner].border = Border(left=left, bottom=bottom)
        for i in range(1, len_N):  # Moving by numbers
            cell = letter + str(int(number) - i)
            ws[cell].border = Border(left=left)
        for i in range(1, len_L):  # Moving by letters
            cell = chr(ord(letter) + i) + number
            ws[cell].border = Border(bottom=bottom)
    elif L == 'BL':
        ws[corner].border = Border(right=right, top=top)
        for i in range(1, len_N):  # Moving by numbers
            cell = letter + str(int(number) + i)
            ws[cell].border = Border(right=right)
        for i in range(1, len_L):  # Moving by letters
            cell = chr(ord(letter) - i) + number
            ws[cell].border = Border(top=top)
    else:
        ws[corner].border = Border(left=left, top=top)
        for i in range(1, len_N):  # Moving by numbers
            cell = letter + str(int(number) + i)
            ws[cell].border = Border(left=left)
        for i in range(1, len_L):  # Moving by letters
            cell = chr(ord(letter) + i) + number
            ws[cell].border = Border(top=top)

    wb.save('18.xlsx')


def top_left_hard_solution(table, a_o_l, corner_N, corner_L, len_N, len_L):
    def resolving(amount_of_lines, return_table, main_table, value):
        for f in range(1, amount_of_lines):
            for j in range(1, amount_of_lines):
                if return_table[f][j] != value:
                    return_table[f][j] = max(return_table[f - 1][j], return_table[f][j - 1]) + main_table[f][j]
        return return_table

    def nullifies(corner_n, len_n, corner_l, len_l, solu_tab, what):
        for NUMBER in range(corner_n - len_n, corner_n):
            for LETTER in range(corner_l - len_l + 1, corner_l + 1):
                solu_tab[NUMBER][LETTER] = what
        return solu_tab

    sum_table_1 = 1
    for i in table:
        sum_table_1 += sum(i)

    sol_tab = create_main_table(a_o_l)  # Table with full solution of this problem like a solution table[NUMBER][LETTER]

    sol_tab[0][0] = table[0][0]
    for i in range(1, a_o_l):
        sol_tab[i][0] = sol_tab[i - 1][0] + table[i][0]
        sol_tab[0][i] = sol_tab[0][i - 1] + table[0][i]

    sol_tab = nullifies(corner_N, len_N, corner_L, len_L, sol_tab, -1)
    sol_tab = resolving(a_o_l, sol_tab, table, -1)
    maximum = sol_tab[a_o_l - 1][a_o_l - 1]  # First answer in our problem

    sol_tab = nullifies(corner_N, len_N, corner_L, len_L, sol_tab, sum_table_1)
    sol_tab = resolving(a_o_l, sol_tab, table, sum_table_1)
    minimum = sol_tab[a_o_l - 1][a_o_l - 1]  # Second answer in our problem
    answer = [maximum, minimum]  # Full answer in one variable(probably changed)

    return answer


def top_left_hard():
    table, amount_of_lines = gen_file()  # Excel table, amount of lines and values in lines
    corner_N, corner_L = random.randint(5, amount_of_lines - 5), random.randint(5, amount_of_lines - 5)
    len_N, len_L = random.randint(2, 4), random.randint(2, 4)
    gen_borders('TL', corner_N, corner_L, len_N, len_L)
    correct_answer = top_left_hard_solution(table, amount_of_lines, corner_N, corner_L, len_N, len_L)
    start(amount_of_lines, 'вниз', 'право', 'нижнюю', 'правую')
    print(
        'Посетив клетку, Робот забирает монету с собой;'
        ' это также относится к начальной и конечной клетке маршрута Робота.')
    print(
        'Откройте файл. Определите максимальную и минимальную денежную сумму,'
        ' которую может собрать Робот, пройдя из левой верхней клетки в правую нижнюю.')
    print(
        'В ответ запишите два числа друг за другом через пробел — сначала максимальную сумму, затем минимальную.')
    print(
        'Исходные данные представляют собой электронную таблицу размером {}*{},'
        ' каждая ячейка которой соответствует клетке квадрата.'.format(
            amount_of_lines, amount_of_lines))
    example()
    print('Для указанных входных данных ответом должна быть пара чисел 41 и 22.')
    return correct_answer


def top_right_hard_solution(table, a_o_l, corner_N, corner_L, len_N, len_L):
    answer = []

    def nullifies(corner_n, len_n, corner_l, len_l, solu_tab, what):
        for number in range(corner_n - len_n, corner_n):
            for letter in range(corner_l, corner_l + len_l):
                solu_tab[number][letter] = what
        return solu_tab

    def resolving(amount_of_lines, return_table, main_table, value):
        for number in range(1, amount_of_lines):
            for letter in range(1, amount_of_lines):
                n_l = amount_of_lines - 1 - letter  # new letter
                n = number
                if return_table[n][n_l] != value:
                    return_table[n][n_l] = max(return_table[n - 1][n_l], return_table[n][n_l + 1]) + main_table[n][n_l]
                else:
                    continue
        return return_table

    sum_table = 1
    for i in table:
        sum_table += sum(i)
    sol_tab = create_main_table(a_o_l)  # solution table two-dimensional massif [NUMBER][LETTER]

    sol_tab[0][-1] = table[0][-1]
    for NUMBER in range(1, a_o_l):
        sol_tab[NUMBER][-1] = sol_tab[NUMBER - 1][-1] + table[NUMBER][-1]
        sol_tab[0][a_o_l - NUMBER - 1] = sol_tab[0][a_o_l - NUMBER] + table[0][
            a_o_l - NUMBER - 1]
    # There is NUMBER like a LETTER
    # I wanted to unite two cycles FOR

    sol_tab = nullifies(corner_N, len_N, corner_L + 1, len_L, sol_tab, -1)
    sol_tab = resolving(a_o_l, sol_tab, table, -1)

    answer.append(sol_tab[-1][0])

    sol_tab = nullifies(corner_N, len_N, corner_L + 1, len_L, sol_tab, sum_table)
    sol_tab = resolving(a_o_l, sol_tab, table, sum_table)

    answer.append(sol_tab[-1][0])
    return answer


def top_right_hard():
    table, amount_of_lines = gen_file()  # Excel table, amount of lines and values in lines
    corner_N, corner_L = random.randint(5, amount_of_lines - 5), random.randint(5, amount_of_lines - 5)
    len_N, len_L = random.randint(2, 4), random.randint(2, 4)
    gen_borders('TR', corner_N, corner_L, len_N, len_L)
    correct_answer = top_right_hard_solution(table, amount_of_lines, corner_N, corner_L, len_N, len_L)
    start(amount_of_lines, 'в низ', 'влево', 'нижнию', 'левую')
    print(
        'Посетив клетку, Робот забирает монету с собой; '
        'это также относится к начальной и конечной клетке маршрута Робота.')
    print(
        'Откройте файл. Определите максимальную и минимальную денежную сумму,'
        ' которую может собрать Робот, пройдя из правой верхней клетки в левую нижнюю.')
    print(
        'В ответ запишите два числа друг за другом через пробел — сначала максимальную сумму, затем минимальную.')
    print(
        'Исходные данные представляют собой электронную таблицу размером {}*{},'
        ' каждая ячейка которой соответствует клетке квадрата.'.format(
            amount_of_lines, amount_of_lines))
    example()
    print('Для указанных входных данных ответом должна быть пара чисел 35 и 15.')
    return correct_answer


def bottom_left_hard_solution(table, a_o_l, corner_N, corner_L, len_N, len_L):
    answer = []

    def nullifies(corner_n, len_n, corner_l, len_l, sol_tab, what):
        for number in range(corner_n - 1, corner_n + len_n - 1):
            for letter in range(corner_n - len_l + 1, corner_l + 1):
                sol_tab[number][letter] = what
        return sol_tab

    def resolving(amount_of_lines, return_table, main_table, value):
        for number in range(1, amount_of_lines):
            for letter in range(1, amount_of_lines):
                n_n = amount_of_lines - number - 1
                # New number a_o_l == 11; NUMBER == 1; => n_n = 9. Like a penult massif
                k = letter
                if return_table[n_n][k] != value:
                    return_table[n_n][k] = max(return_table[n_n][k - 1], return_table[n_n + 1][k]) + main_table[n_n][k]
                else:
                    continue
        return sol_table

    sum_table = 1
    for i in table:
        sum_table += sum(i)
    sol_table = create_main_table(a_o_l)  # solution table two-dimensional massif [NUMBER][LETTER]

    sol_table[-1][0] = table[-1][0]

    for smth in range(1, a_o_l):
        sol_table[-1][smth] = sol_table[-1][smth - 1] + table[-1][smth]
        sol_table[a_o_l - smth - 1][0] = sol_table[a_o_l - smth][0] + table[a_o_l - smth - 1][0]

    sol_table = nullifies(corner_N, len_N, corner_L, len_L, sol_table, -1)
    sol_table = resolving(a_o_l, sol_table, table, -1)

    answer.append(sol_table[0][-1])

    sol_table = nullifies(corner_N, len_N, corner_L, len_L, sol_table, sum_table)
    sol_table = resolving(a_o_l, sol_table, table, sum_table)

    answer.append(sol_table[0][-1])

    return answer


def bottom_left_hard():
    table, amount_of_lines = gen_file()  # Excel table, amount of lines and values in lines
    corner_N, corner_L = random.randint(5, amount_of_lines - 5), random.randint(5, amount_of_lines - 5)
    len_N, len_L = random.randint(2, 4), random.randint(2, 4)
    gen_borders('BL', corner_N, corner_L, len_N, len_L)
    correct_answer = bottom_left_hard_solution(table, amount_of_lines, corner_N, corner_L, len_N, len_L)
    start(amount_of_lines, 'вверх', 'право', 'верхнюю', 'правую')
    print(
        'Посетив клетку, Робот забирает монету с собой;'
        ' это также относится к начальной и конечной клетке маршрута Робота.')
    print(
        'Откройте файл. Определите максимальную и минимальную денежную сумму,'
        ' которую может собрать Робот, пройдя из левой нижней клетки в правую верхнюю.')
    print(
        'В ответ запишите два числа друг за другом через пробел — сначала максимальную сумму, затем минимальную.')
    print(
        'Исходные данные представляют собой электронную таблицу размером {}*{},'
        ' каждая ячейка которой соответствует клетке квадрата.'.format(
            amount_of_lines, amount_of_lines))
    example()
    print('Для указанных входных данных ответом должна быть пара чисел 35 и 15.')
    return correct_answer


def bottom_right_hard_solution(table, a_o_l, corner_N, corner_L, len_N, len_L):
    answer = []

    def nullifies(corner_n, len_n, corner_l, len_l, sol_tab, what):
        for number in range(corner_n - 1, corner_n + len_n - 1):
            for letter in range(corner_l, corner_l + len_l):
                sol_tab[number][letter] = what
        return sol_tab

    def resolving(amount_of_lines, return_table, main_table, value):
        for number in range(1, amount_of_lines):
            for letter in range(1, amount_of_lines):
                n_n = amount_of_lines - number - 1
                # New number amount_of_lines == 11; NUMBER == 1; => n_n = 9. Like a penult massif
                k = amount_of_lines - letter - 1
                if return_table[n_n][k] != value:
                    return_table[n_n][k] = max(return_table[n_n][k + 1], return_table[n_n + 1][k]) + main_table[n_n][k]
                else:
                    continue
        return return_table

    sol_table = create_main_table(a_o_l)  # solution table two-dimensional massif [NUMBER][LETTER]
    sum_table = 1
    for i in table:
        sum_table += sum(i)

    sol_table[-1][-1] = table[-1][-1]

    for smth in range(1, a_o_l):
        sol_table[-1][a_o_l - 1 - smth] = sol_table[-1][a_o_l - smth] + table[-1][a_o_l - 1 - smth]
        sol_table[a_o_l - smth - 1][-1] = sol_table[a_o_l - smth][-1] + table[a_o_l - smth - 1][-1]

    sol_table = nullifies(corner_N, len_N, corner_L, len_L, sol_table, -1)
    sol_table = resolving(a_o_l, sol_table, table, -1)

    answer.append(sol_table[0][0])

    sol_table = nullifies(corner_N, len_N, corner_L, len_L, sol_table, sum_table)
    sol_table = resolving(a_o_l, sol_table, table, sum_table)

    answer.append(sol_table[0][0])
    return answer


def bottom_right_hard():
    table, amount_of_lines = gen_file()  # Excel table, amount of lines and values in lines
    corner_N, corner_L = random.randint(5, amount_of_lines - 5), random.randint(5, amount_of_lines - 5)
    len_N, len_L = random.randint(2, 4), random.randint(2, 4)
    gen_borders('BR', corner_N, corner_L, len_N, len_L)

    correct_answer = bottom_right_hard_solution(table, amount_of_lines, corner_N, corner_L, len_N, len_L)
    start(amount_of_lines, 'вверх', 'влево', 'верхнюю', 'левую')
    print(
        'Посетив клетку, Робот забирает монету с собой; '
        'это также относится к начальной и конечной клетке маршрута Робота.')
    print(
        'Откройте файл. Определите максимальную и минимальную денежную сумму,'
        ' которую может собрать Робот, пройдя из правой нижней клетки в левую верхнюю.')
    print(
        'В ответ запишите два числа друг за другом через пробел — сначала максимальную сумму, затем минимальную.')
    print(
        'Исходные данные представляют собой электронную таблицу размером {}*{},'
        ' каждая ячейка которой соответствует клетке квадрата.'.format(
            amount_of_lines, amount_of_lines))
    example()
    print('Для указанных входных данных ответом должна быть пара чисел 41 и 22.')
    return correct_answer
